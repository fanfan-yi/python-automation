import cx_Oracle
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.styles import Font, PatternFill
from datetime import datetime

# ===== Oracle 連線 =====
cx_Oracle.init_oracle_client(lib_dir=r"D:\instantclient_21_7")#Oracle 驅動位置

dsn = cx_Oracle.makedsn("192.168.100.43", 1541, service_name="XXX")#makedsn(host, port, service_name)封裝 Oracle 資料庫連線參數
conn = cx_Oracle.connect(user="XXX", password="XXX", dsn=dsn)

# ===== SQL =====
sql_daily_po = """
SELECT
    TRUNC(creation_date) AS po_date,
    COUNT(*) AS po_count
FROM po_headers_all
WHERE creation_date >= SYSDATE - 30
GROUP BY TRUNC(creation_date)
ORDER BY po_date
"""

sql_po_status = """
SELECT
    authorization_status AS po_status,
    COUNT(*) AS po_count
FROM po_headers_all
WHERE creation_date >= SYSDATE - 30
GROUP BY authorization_status
ORDER BY authorization_status
"""

# ===== 撈資料 =====
df_daily = pd.read_sql(sql_daily_po, conn)
df_status = pd.read_sql(sql_po_status, conn)

conn.close()

# ===== 輸出 Excel =====
file_name = f"PO_Report_{datetime.now().strftime('%Y%m%d')}.xlsx"

with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
    df_daily.to_excel(writer, sheet_name="PO_Daily_Trend", index=False)
    df_status.to_excel(writer, sheet_name="PO_Status", index=False)

# ===== Excel 畫圖 =====
wb = load_workbook(file_name)#用 openpyxl 打開 Excel畫圖

# ---------- PO 趨勢圖 ----------
ws_trend = wb["PO_Daily_Trend"]#點進工作表:PO_Daily_Trend

line_chart = LineChart()
line_chart.title = "Daily PO Creation Trend"
line_chart.y_axis.title = "PO Count"
line_chart.x_axis.title = "Date"
#---------- 指定資料範圍 ----------
data_ref = Reference(ws_trend, min_col=2, min_row=1, max_row=ws_trend.max_row)#數據範圍B1-MAX
cats_ref = Reference(ws_trend, min_col=1, min_row=2, max_row=ws_trend.max_row)#數據範圍A2-MAX

line_chart.add_data(data_ref, titles_from_data=True)
line_chart.set_categories(cats_ref)

ws_trend.add_chart(line_chart, "E2")#圖表開始位置:E2

# ---------- PO 狀態分布 ----------
ws_status = wb["PO_Status"]

bar_chart = BarChart()
bar_chart.title = "PO Status Distribution"
bar_chart.y_axis.title = "PO Count"
bar_chart.x_axis.title = "Status"

data_ref = Reference(ws_status, min_col=2, min_row=1, max_row=ws_status.max_row)
cats_ref = Reference(ws_status, min_col=1, min_row=2, max_row=ws_status.max_row)

bar_chart.add_data(data_ref, titles_from_data=True)
bar_chart.set_categories(cats_ref)

ws_status.add_chart(bar_chart, "E2")

# ---------- 表頭加粗 ----------
#每一張 sheet
for sheet in wb.worksheets:
    #sheet[1] = 第 1 列
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.font = Font(size = 14)
        cell.fill = PatternFill(start_color="DDDDDD", fill_type="solid")

wb.save(file_name)

print(f"PO Excel report generated: {file_name}")

